"""
check the judgement
"""

import os
import json
import sys
import re
import argparse
import codecs
from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("judged_file")
    args=parser.parse_args()

    wb = load_workbook(args.judged_file)
    ws = wb.get_sheet_by_name('Sheet1')
    judge = {}
    printed = {}
    for row in ws.iter_rows():
        entity = row[1].value
        sentence = row[2].value
        judgement = row[3].value
        if sentence not in judge:
            judge[sentence] = judgement
        else:
            if judgement != judge[sentence]:
                if sentence not in printed:
                    printed[sentence] = 1
                    
                    print sentence


if __name__=="__main__":
    main()

