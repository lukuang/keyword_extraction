"""
combine judgment with the newly generate no-redundant xlsx file
"""

import os
import json
import sys
import re
import argparse
import codecs
from openpyxl import load_workbook
import xlsxwriter

def get_judged_data(judged_file):
    """Return judged data for xlsx file as 
    {entity:{sentence:judgement} } 
    """

    judged_data = {}
    wb = load_workbook(judged_file)
    ws = wb.get_sheet_by_name('Sheet1')
    for row in ws.iter_rows():
        entity = row[1].value
        sentence = row[2].value
        judgement = row[3].value
        if entity not in judged_data:
            judged_data[entity] = {}

        judged_data[entity][sentence] = judgement
        


    return judged_data

def merge(judged_data,no_redundant_file,dest_file):
    wb = load_workbook(no_redundant_file)
    ws = wb.get_sheet_by_name('Sheet1')

    new_data = []
    for row in ws.iter_rows():
        instance = row[0].value
        entity = row[1].value
        sentence = row[2].value

        single_data = {
            "instance":instance,
            "sentence":sentence,
            "entity":entity
        }
        if entity in judged_data:
            if sentence in judged_data[entity]:
                single_data["judgement"] = judged_data[entity][sentence]

        new_data.append(single_data)
    
    
    # write new data
    workbook = xlsxwriter.Workbook(dest_file)
    worksheet = workbook.add_worksheet()

    row = 0
    col = 0

    for single_data in new_data:
        instance = single_data["instance"]
        entity = single_data["entity"]
        sentence = single_data["sentence"]

        worksheet.write(row, col,     instance)
        worksheet.write(row, col + 1, entity)
        worksheet.write(row, col + 2, sentence)

        if "judgement" in single_data:
            worksheet.write(row, col+3,single_data["judgement"])
        row += 1

    workbook.close()


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("judged_file")
    parser.add_argument("no_redundant_file")
    parser.add_argument("dest_file")
    args=parser.parse_args()

    judged_data =  get_judged_data(args.judged_file)
    merge(judged_data,args.no_redundant_file,args.dest_file)


if __name__=="__main__":
    main()

