# -*- coding: utf-8 -*-
"""
use judged data to create basic data used to generate training/testing data
"""

import os
import json
import sys
import re
import argparse
import codecs
from openpyxl import load_workbook
from myUtility.corpus import Sentence, Document, Model
from get_entity_cate import get_cate_for_entity_list
reload(sys)
sys.setdefaultencoding("utf-8")


def get_cate_info(pure_entities,cate_info_file):
    if os.path.exists(cate_info_file):
        cate_info = json.load(open(cate_info_file))
        new_cate = []
        for entity in pure_entities:
            if entity not in cate_info:
                new_cate.append(entity)
        if len(new_cate)!=0:
            cate_info.update(get_cate_for_entity_list(new_cate))
        with codecs.open(cate_info_file,'w','utf-8') as f:
            f.write(json.dumps(cate_info)) 
    else:
        cate_info = get_cate_for_entity_list(list(pure_entities) )
        with codecs.open(cate_info_file,'w','utf-8') as f:
            f.write(json.dumps(cate_info)) 
    return cate_info

def get_type_info(candidate_file,required_types):
    type_info = {}

    for t in required_types:
        type_info[t] = set()

    with open(candidate_file) as f:
        for line in f:
            line = line.rstrip()
            m = re.search("^([A-Z]+):(.+?):(.+)$",line)
            #parts = line.split(':')

            try:
                entity_type = m.group(1)
            except Exception:
                print line
                sys.exit(-1)
            if entity_type in required_types:
                entity = unicode(m.group(2))
                #if entity == u"Kosair Children’s Hospital":
                #    print "Found sentence %s" %line
                type_info[entity_type].add(entity)

    return type_info




def get_judged_data(source_file,type_info,cate_info_file):
    """Return judged data for xlsx file as 
    [{instance,entity,sentence,judgement,type,category}]
    """

    entities = set()
    judged_data = []
    wb = load_workbook(source_file)
    ws = wb.get_sheet_by_name('Sheet1')
    for row in ws.iter_rows():
        instance = row[0].value
        entity = unicode(row[1].value)
        sentence = row[2].value
        judgement = row[3].value

        types = []
        found_type = False
        for entity_type in type_info:
            if entity in type_info[entity_type]:
                types.append(entity_type)
                found_type = True
        if not found_type:
            print "no type for %s" %(entity)
            sys.exit(-1)

        entities.add(entity)
        single_data = {
                "instance":instance,
                "judgement":judgement,
                "sentence":sentence,
                "entity":entity,
                "type":types
            }
        judged_data.append(single_data)

    cate_info = get_cate_info(list(entities),cate_info_file)

    for single_data in judged_data:
        entity = single_data["entity"]
        single_data["category"] = cate_info[entity]


    return judged_data




def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source_file")
    parser.add_argument("candidate_file")
    parser.add_argument("dest_file")
    parser.add_argument("--required_types", "-rt",nargs='+',default=["ORGANIZATION","LOCATION"])
    parser.add_argument("--cate_info_file","-cf",default="/home/1546/code/keyword_extraction/sentence_level/data/other_data/cate_info.json")
    args=parser.parse_args()

    type_info = get_type_info(args.candidate_file,args.required_types)

    judged_data = get_judged_data(args.source_file,type_info,args.cate_info_file)

    with codecs.open(args.dest_file,"w","utf-8") as f:
        f.write(json.dumps(judged_data))


if __name__=="__main__":
    main()

